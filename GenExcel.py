#coding=utf-8

import xlrd 
import xlwt 
import pickle
import xlutils
from xlutils.copy import copy
import openpyxl
from openpyxl.styles import Border, Side, Font
from collections import namedtuple
#import pandas
#import struct
#from xlwt import *
#f = open("D:\demo1.xlsx", 'w') 创建Excel文件不要用open，而应该使用包xlwt下的API
#f.close()
#workbook = xlrd.open_workbook("D:\demo1.xlsx") 



#wb = xlwt.Workbook()
#ws = wb.add_sheet("sheetwt")#Name sheet.Default formate:Sheet1、Sheet2...
#ws.write(3, 1, "wow")#Row and Column counted from 0.
##wb.save("D:\\tmp.xlsx")#use \\ not \
#wb.save("//tool//tmp.xlsx")



#excel ---->  101010100101
#101010100101  ----->  excel

##############################################构建内容##############################################
######构造一个结构体，把要填充的内容以及位置存到结构体。众多这样的结构体构成一个tuple，写Excel时去

#PCTuple = (0, 1, "content")

#PCTuple = namedtuple(row, col, content)
PCTuple = namedtuple('PCTuple', 'row col content')

class OperExcel:
    
    def GenTemplate(self, type):
        pass
    
    def RepairCell(self, path):
        pass
    
    def FillContent(self, path, pctl):
        wb = openpyxl.load_workbook(path)
        sheet_names = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet_names[0])
        
        print(type(ws))
        print(ws)
        
        #ws.cell(row = 14, column = 8).value = "test"
        for i in pctl:
            ws.cell(row = i.row, column = i.col).value = i.content
        
        #修复边框合并单元格
        border=Border(left=Side(border_style='thin', color='FF000000'),
                      right=Side(border_style='thin', color='FF000000'),
                            top=Side(border_style='thin', color='FF000000'),
                            bottom=Side(border_style='thin', color='FF000000'))      
        #openpyxl.cell.cell.Cell.
        for row in ws.iter_rows('A1:K12'):
            for c in row:
                c.border = border
                
       
        ws.merge_cells('A8:K8')
        ws.merge_cells('A12:K12')
        wb.save(path)
        #xrd = xlrd.open_workbook(path)
        #wksheet = xrd.sheets()[0]
        #nrow = wksheet.nrows
        #ncol = wksheet.ncols
        #print(nrow)
        #print(ncol)
        #wksheet.write(13, 7, content)
        #xwb = copy(xrd)
        #ws = xwb.get_sheet(0)
        #ws.write(13, 7, "test")
        #xwb.save(path)
    
    def GetExcelInfo(self, path):
        f = open("D:\\tmpdump", 'w')
        rd = xlrd.open_workbook(path)
        rd.dump(f)
        f.close()   
        
    def PackExcel(self, path):
        #rd = xlrd.open_workbook(path)
        #pkf = open("D:\\tmp2.pkl", 'wb')
        #print(type(rd))
        #pickle.dump(rd, pkf)
        #pkf.close()
        #wt = xlwt.Workbook(path)
        rd = xlrd.open_workbook(path)
        wb = copy(rd)
        pkf = open("D:\\tmp2.pkl", 'wb')
        print(type(wb))
        pickle.dump(wb, pkf)
        pkf.close()        
    
    def UnpackExcel(self, path):
        pkf = open(path, 'rb')
        print(type(pkf))
        pk = pickle.load(pkf)
        print(type(pk))
        pk.save("D:\\tmp2.xlsx")
    
if __name__ == "__main__":
    pass
    #oe = OperExcel()
    ##oe.PackExcel("D:\\tmp.xlsx")
    ##oe.UnpackExcel("D:\\tmp2.pkl")
    ##oe.FillContent("C:\\Users\\Henrry\\Desktop\\apache.xlsx", 0, 0, "content")
    #PCTuple = namedtuple('PCTuple', 'row col content')
    
    
    
    #pct1 = PCTuple(14, 8, "test1")
    #pct2 = PCTuple(15, 8, "test2")
    #pct3 = PCTuple(16, 8, "test3")
    
    
    
    #pctl = []
    #pctl.append(pct1)
    #pctl.append(pct2)
    #pctl.append(pct3)
    #oe.FillContent("C:\\Users\\Henrry\\Desktop\\apache.xlsx", pctl)